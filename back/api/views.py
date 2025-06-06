
from rest_framework import viewsets
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.permissions import IsAuthenticated
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from django.db.models import Subquery, OuterRef
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import openpyxl

from .models import *
from .serializers import *


# CRUD de Sensores
class SensorListCreateView(ListCreateAPIView):
    # permission_classes = [IsAuthenticated]
    queryset = Sensor.objects.all()
    serializer_class = SensorSerializer

class SensorDetailView(RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Sensor.objects.all()
    serializer_class = SensorSerializer


# CRUD de Ambientes
class AmbienteViewSet(ListCreateAPIView):
    queryset = Ambiente.objects.all()
    serializer_class = AmbienteSerializer


# CRUD de Histórico com filtro
class HistoricoViewSet(ListCreateAPIView):
    queryset = Historico.objects.all()
    serializer_class = HistoricoSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['sensor', 'ambiente', 'timestamp']

    def get_queryset(self):
        queryset = self.queryset
        request = self.request
        if request.method == "GET":
            sensores = Sensor.objects.filter(sensor__iexact=request.GET.get("tipo")).values("id")

            return queryset.filter(id__in=Subquery(sensores))

        return queryset


# Importação via planilha Excel
@csrf_exempt
def upload(request):
    if request.method == 'POST':
        tipo = request.POST.get("tipo")
        excel_file = request.FILES.get("arquivo")

        if not excel_file:
            return JsonResponse({"erro": "Nenhum arquivo enviado."}, status=400)

        if not excel_file.name.endswith('.xlsx'):
            return JsonResponse({"erro": "O arquivo deve ter a extensão .xlsx"}, status=400)

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            if tipo == "ambiente":
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    Ambiente.objects.create(
                        sig=row[0],
                        descricao=row[1],
                        ni=row[2],
                        responsavel=row[3]
                    )
                return JsonResponse({"status": "Ambientes importados com sucesso."})

            elif tipo == "sensor":
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    Sensor.objects.create(
                        sensor=row[0],
                        mac_address=row[1],
                        unidade_med=row[2],
                        latitude=float(row[3]),
                        longitude=float(row[4]),
                        status=row[5] == True or row[5] == "Ativo"
                    )
                return JsonResponse({"status": "Sensores importados com sucesso."})

            elif tipo == "historico":
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    Historico.objects.create(
                        sensor_id=int(row[0]),
                        ambiente_id=int(row[1]),
                        valor=str(row[2])
                    )
                return JsonResponse({"status": "Histórico importado com sucesso."})

            else:
                return JsonResponse({"erro": "Tipo inválido. Use 'ambiente', 'sensor' ou 'historico'."}, status=400)

        except Exception as e:
            return JsonResponse({"erro": f"Erro ao processar planilha: {str(e)}"}, status=500)

    return JsonResponse({"erro": "Método não permitido. Use POST."}, status=405)


# View para dados de gráfico por tipo de medida
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def grafico_medida(request, tipo):
    """
    Retorna os últimos 20 dados do histórico para a medida informada (temperatura, umidade, etc.)
    """
    historicos = Historico.objects.filter(sensor__unidade_med=tipo).order_by('-timestamp')[:20]

    dados = [
        {
            "timestamp": h.timestamp.strftime("%H:%M"),
            "valor": float(h.valor)
        }
        for h in reversed(historicos)  # ordem cronológica
    ]

    return Response(dados)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def historico_leitura(request):
    historicos = (Historico.objects.all())
    hist = HistoricoSerializer(historicos)

    return JsonResponse(hist.data)
